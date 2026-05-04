from pathlib import Path
import shutil

from openpyxl import load_workbook

from app.models.domain import VocabRow
from app.services.workbook import DEFAULT_TEMPLATE_NAME, HIGHLIGHT_FONT_SIZE_INCREMENT, WorkbookService


def test_fill_template_writes_expected_columns_and_highlights_example_term() -> None:
    root = Path("D:/workspace/vocab-sheet-service/.runtime/test-workbook")
    if root.exists():
        shutil.rmtree(root, ignore_errors=True)
    root.mkdir(parents=True, exist_ok=True)
    output = root / "result.xlsx"
    rows = [
        VocabRow(
            word="apple",
            ipa="/aepl/",
            pos_abbr="n.",
            zh_meaning="苹果",
            example="This apple is red. Apple pies are great.",
            example_page=12,
            sources=["pdf:qwen_vl", "manual"],
        )
    ]

    service = WorkbookService()
    assert service.template_path.name == DEFAULT_TEMPLATE_NAME
    template_workbook = load_workbook(service.template_path)
    try:
        expected_highlight_size = (template_workbook["Sheet1"]["F3"].font.sz or 11) + HIGHLIGHT_FONT_SIZE_INCREMENT
    finally:
        template_workbook.close()

    service.fill_template(rows, output, workbook_title="lesson")

    workbook = load_workbook(output, rich_text=True)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A1"].value == "lesson"
        assert sheet["A3"].value == 1
        assert sheet["B3"].value == "apple"
        assert sheet["C3"].value == "/aepl/"
        assert sheet["D3"].value == "n."
        assert sheet["E3"].value == "苹果"
        assert str(sheet["F3"].value) == "This apple is red. Apple pies are great."
        highlighted_segments = [
            block.text
            for block in sheet["F3"].value
            if hasattr(block, "text") and getattr(getattr(block, "font", None), "b", None)
        ]
        assert highlighted_segments == ["apple", "Apple"]
        highlighted_font_sizes = [
            block.font.sz
            for block in sheet["F3"].value
            if hasattr(block, "text") and getattr(getattr(block, "font", None), "b", None)
        ]
        assert highlighted_font_sizes == [expected_highlight_size, expected_highlight_size]
        assert sheet["G3"].value == 12
        assert sheet["H3"].value == "教材 / 手工补词"
    finally:
        workbook.close()
        shutil.rmtree(root, ignore_errors=True)


def test_fill_template_keeps_blank_example_for_manual_only_row() -> None:
    root = Path("D:/workspace/vocab-sheet-service/.runtime/test-workbook-plain")
    if root.exists():
        shutil.rmtree(root, ignore_errors=True)
    root.mkdir(parents=True, exist_ok=True)
    output = root / "result.xlsx"
    rows = [
        VocabRow(
            word="idea",
            ipa="/aɪˈdɪə/",
            pos_abbr="n.",
            zh_meaning="想法",
            example="",
            example_page=None,
            sources=["manual"],
        )
    ]

    WorkbookService().fill_template(rows, output, workbook_title="plain-lesson")

    workbook = load_workbook(output, rich_text=True)
    try:
        sheet = workbook["Sheet1"]
        assert sheet["A1"].value == "plain-lesson"
        assert sheet["F3"].value in ("", None)
        assert sheet["G3"].value is None
        assert sheet["H3"].value == "手工补词"
    finally:
        workbook.close()
        shutil.rmtree(root, ignore_errors=True)
