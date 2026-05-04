from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from app.models.domain import VocabRow
from app.utils.paths import resource_path
from app.utils.text import find_term_spans


EXPECTED_HEADERS = ["序号", "单词", "音标", "词性", "中文意思", "例句", "例句在教材中出现的页数", "来源"]
DEFAULT_TEMPLATE_NAME = "单词表模板.xlsx"
DEFAULT_TEMPLATE_PATH = resource_path("template", DEFAULT_TEMPLATE_NAME)
TEMPLATE_TITLE_RULE = "使用上传教材文件名（不含扩展名）作为第一行标题"
TEMPLATE_COLUMNS = [
    {"key": "index", "label": "序号"},
    {"key": "word", "label": "单词"},
    {"key": "ipa", "label": "音标"},
    {"key": "pos_abbr", "label": "词性"},
    {"key": "zh_meaning", "label": "中文意思"},
    {"key": "example", "label": "例句"},
    {"key": "example_page", "label": "例句在教材中出现的页数"},
    {"key": "source", "label": "来源"},
]
TEMPLATE_RESULT_FIELDS = [
    {"key": "output_filename", "label": "下载文件名", "description": "生成后的 Excel 文件名"},
    {"key": "rows_written", "label": "写入词数", "description": "成功写入模板的词条数量"},
    {"key": "skipped_count", "label": "跳过词数", "description": "未写入模板的词条数量"},
    {"key": "written_rows", "label": "写入明细", "description": "每个已写入词条的列值与来源"},
    {"key": "skipped_items", "label": "跳过明细", "description": "每个跳过词条的原因与来源"},
]
HIGHLIGHT_FONT_SIZE_INCREMENT = 2


class WorkbookTemplateError(RuntimeError):
    """Raised when the fixed workbook template is unavailable or invalid."""


class WorkbookService:
    def __init__(self, template_path: Path | None = None) -> None:
        self.template_path = template_path or DEFAULT_TEMPLATE_PATH

    def fill_template(self, rows: list[VocabRow], output_path: Path, *, workbook_title: str | None = None) -> None:
        if not self.template_path.exists():
            raise WorkbookTemplateError(f"固定模板不存在: {self.template_path}")

        workbook = load_workbook(self.template_path)
        try:
            if "Sheet1" not in workbook.sheetnames:
                raise WorkbookTemplateError("固定模板缺少 Sheet1 工作表。")

            sheet = workbook["Sheet1"]
            self._validate_headers(sheet)
            self._write_title(sheet, workbook_title)
            self._clear_existing_rows(sheet)

            for index, row in enumerate(rows, start=3):
                sheet.cell(row=index, column=1, value=index - 2)
                sheet.cell(row=index, column=2, value=row.word)
                sheet.cell(row=index, column=3, value=row.ipa)
                sheet.cell(row=index, column=4, value=row.pos_abbr)
                sheet.cell(row=index, column=5, value=row.zh_meaning)
                example_cell = sheet.cell(row=index, column=6)
                example_cell.value = (
                    _build_example_value(row.example, row.word, base_font_size=example_cell.font.sz)
                    if row.example
                    else ""
                )
                sheet.cell(row=index, column=7, value=row.example_page)
                sheet.cell(row=index, column=8, value=_format_source_value(row.sources))

                alignment = copy(example_cell.alignment)
                alignment.wrap_text = True
                example_cell.alignment = alignment

            workbook.save(output_path)
        finally:
            workbook.close()

    def _validate_headers(self, sheet) -> None:
        actual = [sheet.cell(row=2, column=index).value for index in range(1, 9)]
        if actual != EXPECTED_HEADERS:
            raise WorkbookTemplateError(
                f"固定模板表头不符合预期。第2行应为 {EXPECTED_HEADERS}，实际为 {actual}"
            )

    def _write_title(self, sheet, workbook_title: str | None) -> None:
        title = (workbook_title or "").strip()
        if title:
            sheet["A1"] = title

    def _clear_existing_rows(self, sheet) -> None:
        for row_index in range(3, sheet.max_row + 1):
            for column_index in range(1, 9):
                sheet.cell(row=row_index, column=column_index, value=None)


def _build_example_value(example: str, term: str, *, base_font_size: float | None = None) -> str | CellRichText:
    spans = find_term_spans(example, term)
    if not spans:
        return example

    parts: list[str | TextBlock] = []
    cursor = 0
    for start, end in spans:
        if start > cursor:
            parts.append(example[cursor:start])
        parts.append(TextBlock(_build_highlight_font(base_font_size), example[start:end]))
        cursor = end
    if cursor < len(example):
        parts.append(example[cursor:])
    return CellRichText(*parts)


def _build_highlight_font(base_font_size: float | None) -> InlineFont:
    font_size = (base_font_size or 11) + HIGHLIGHT_FONT_SIZE_INCREMENT
    return InlineFont(b=True, color="FFFF0000", sz=font_size)


def _format_source_value(sources: list[str]) -> str:
    labels: list[str] = []
    if any(source == "manual" for source in sources):
        labels.append("手工补词")
    if any(source == "audio" for source in sources):
        labels.append("音频补词")
    if any(source not in {"manual", "audio"} for source in sources) or not labels:
        labels.insert(0, "教材")

    deduped: list[str] = []
    for label in labels:
        if label not in deduped:
            deduped.append(label)
    return " / ".join(deduped)
